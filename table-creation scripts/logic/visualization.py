import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
import os

from openpyxl.utils import get_column_letter


class visualization:

    def __init__(self):
        if os.path.exists('../out/wb.xlsx'):
            os.remove('../out/wb.xlsx')

        self.workbook = Workbook()

        # these course names will appear in the table
        courses = ['BS - Year 1 (Computer Engineering)', 'BS - Year 1 (Computer Engineering)']
        # these group names are used to fetch data for groups
        groups = [['B20-CE-01', 'B20-CE-01', 'B20-CE-01'], ['B20-CE-01', 'B20-CE-01']]
        # these colors are used for courses
        palettes = [{'day': "00FF0000", 'main': "00FFFF00"},
                    {'day': "00FFF000", 'main': "00FF0FF0"}]

        # width of columns
        self.column_width = 20

        # there are this much time slots in a day
        self.time_slots_number = 7

        # apply certain styles for each cell in the table
        self.post_processing()

        # time slot takes this much rows
        self.time_slot_height = 3

        # there will be this much days in a week
        self.week = 5

        # a day takes this much rows
        self.day_height = 1 + self.time_slot_height * self.time_slots_number

        # first create the first column
        self.create_left()

        # each group takes this much columns
        self.group_width = 2

        # then add courses which start from the 2nd column

        # calculate all starting columns for courses
        course_sizes = [0 if i == 0 else len(groups[i - 1]) for i in range(len(groups))]
        for i in range(1, len(course_sizes)):
            course_sizes[i] = course_sizes[i - 1] + course_sizes[i]
        course_columns = [2 + _ * 2 for _ in course_sizes]

        # add each course
        for course_index in range(len(courses)):
            self.add_course(course_name=courses[course_index],
                            groups=groups[course_index],
                            course_column=course_columns[course_index],
                            course_palette=palettes[course_index])

        self.post_processing()

        # save the notebook
        self.workbook.save(filename='../out/wb.xlsx')

    def add_course(self, course_column, course_name, groups, course_palette):

        # width of a course in columns
        course_width = len(groups) * self.group_width

        # add header of a course
        sheet = self.workbook.active

        # add course_name
        sheet.merge_cells(start_row=1, start_column=course_column,
                          end_row=1, end_column=course_column + course_width - 1)
        sheet.cell(row=1, column=course_column).fill = PatternFill("solid", fgColor=course_palette['main'])
        sheet.cell(row=1, column=course_column).value = course_name

        # first row of each day inside the course should look like merged cells
        for day_index in range(self.week):
            # starting row of a day
            day_row = 3 + day_index * self.day_height
            sheet.merge_cells(start_row=day_row,
                              start_column=course_column,
                              end_row=day_row,
                              end_column=course_column + course_width - 1)
            sheet.cell(row=day_row, column=course_column).fill = \
                PatternFill("solid", fgColor=course_palette['day'])

        # add groups
        group_column = course_column
        for group_index in range(len(groups)):
            # all non-empty slots of a course will be painted this color
            group_column = course_column + group_index * self.group_width
            self.add_group(group_column, groups[group_index], course_palette['main'])

        # need to merge all course cells for Lectures and Tutorials
        for row in range(1, 2 + self.week * self.day_height + 1):
            value = str(sheet.cell(row=row, column=course_column).value)
            if value.find('Lec') != -1 or value.find('Tut') != -1:
                for index in range(3):
                    sheet.merge_cells(start_row=row + index, start_column=course_column,
                                      end_row=row + index, end_column=course_column + course_width - 1)

    # adds a column with group data into the table
    def add_group(self, group_column, group_name, color):
        wb = self.workbook
        sheet = wb.active

        # 2 first rows are for course name, and group_name
        group_row = 2

        # put group name into the table
        sheet.merge_cells(start_row=group_row, start_column=group_column,
                          end_row=group_row, end_column=group_column + self.group_width - 1)
        sheet.cell(row=group_row, column=group_column).value = group_name
        sheet.cell(row=group_row, column=group_column).fill = PatternFill("solid", fgColor=color)

        # read data for group
        df = pd.read_csv(f'../groups_data/{group_name}.csv', index_col=0)

        # there are 3 columns of data for each day in a dataframe
        day_data_width = 3

        # iterate through df by days
        for day_index in range(self.week - 1):
            # work with subset of dataframe

            # calculate the starting column for a day
            day_column = day_index * day_data_width

            # select data for one day; no offset since headers are not data
            day_data = df.iloc[0: self.time_slots_number, day_column: day_column + day_data_width]

            # calculate starting row and column for a day
            day_row = group_row + 1 + self.day_height * day_index
            day_column = group_column

            self.add_day(day_row, day_column, day_data, color)

    def add_day(self, day_row, day_column, day_data, color):
        sheet = self.workbook.active

        # iterate through time slots for a day
        for time_slot_index in range(self.time_slots_number):
            # for each day, first row is occupied by an empty row
            time_slot_row = day_row + 1 + time_slot_index * self.time_slot_height

            # put course_name, teacher and room
            # common stands for a lecture or a tutorial
            common = False

            for index in range(3):
                row, column = time_slot_row, day_column

                self.add_border([time_slot_row, day_column, time_slot_row + 2, day_column])

                # no data in the dataframe - just merge all cells in the timeslot
                if pd.isnull(day_data.iloc[time_slot_index, index]):
                    sheet.merge_cells(start_row=row, start_column=column,
                                      end_row=row + 2, end_column=column + 1)
                    continue

                cell = sheet.cell(row=row + index, column=column)
                cell.value = str(day_data.iloc[time_slot_index, index])
                cell.fill = PatternFill("solid", fgColor=color)

                # usually, pairs of cells are merged
                # but, if it is a lecture or a tutorial cell, it shouldn't be merged
                if str(cell.value).find('Lec') != -1 or str(cell.value).find('Tut') != -1:
                    common = True

                if common is False:
                    sheet.merge_cells(start_row=row + index, start_column=column,
                                      end_row=row + index, end_column=column + 1)

    # creates the leftmost column with names of week days and timeslots
    def create_left(self):
        sheet = self.workbook.active

        # fill it with some content
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        time_slots = ["9:00-10:30", "10:40-12:10", "12:40-14:10", "14:20-15:50", "16:00-17:30", "17:40-19:10",
                      "19:20-20:50"]

        # vertical offset
        offset = 3

        # height of a slot in rows
        time_slot_height = 3

        # number of time_slots
        slots_number = len(time_slots)

        # one row for name of day and rows for time_slots
        day_height = 1 + slots_number * time_slot_height
        week = len(days)

        for day_index in range(week):
            day_row = offset + day_index * day_height
            sheet[f'A{day_row}'] = days[day_index]

            for slot_index in range(slots_number):
                slot_row = day_row + 1 + slot_index * 3
                sheet.merge_cells(f'A{slot_row}:A{slot_row + 2}')

                sheet.row_dimensions[slot_row].height = 40
                sheet.row_dimensions[slot_row + 1].height = 40
                sheet.row_dimensions[slot_row + 2].height = 20

                sheet[f'A{slot_row}'] = time_slots[slot_index]
                sheet[f'A{slot_row}'].alignment = Alignment(vertical="top")

    def post_processing(self):
        # styles applied to all cells

        sheet = self.workbook.active

        # alignment
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text='True')

        # width of columns
        for col in sheet.columns:
            sheet.column_dimensions[get_column_letter(col[0].column)].width = self.column_width

        # borders
        side_style = Side(style='thin')
        for row in sheet.iter_rows():
            for cell in row:
                if cell.border is None:
                    cell.border = Border(left=side_style, top=side_style, right=side_style, bottom=side_style)

        sheet.freeze_panes = sheet['B3']

    # add border around some rectangular range of cells
    def add_border(self, cell_range):
        sheet = self.workbook.active
        side_style = Side(style='thin')

        min_row, min_column, max_row, max_column = cell_range
        for row in range(min_row, max_row + 1):
            sheet.cell(row, min_column).border = Border(left=side_style)
            sheet.cell(row, max_column).border = Border(right=side_style)
        for column in range(min_column, max_column + 1):
            sheet.cell(min_row, column).border = Border(top=side_style)
            sheet.cell(max_row, column).border = Border(bottom=side_style)

        sheet.cell(min_row, min_column).border = Border(left=side_style, top=side_style)
        sheet.cell(min_row, max_column).border = Border(right=side_style, top=side_style)
        sheet.cell(max_row, min_column).border = Border(left=side_style, bottom=side_style)
        sheet.cell(max_row, max_column).border = Border(right=side_style, bottom=side_style)


vis = visualization()
