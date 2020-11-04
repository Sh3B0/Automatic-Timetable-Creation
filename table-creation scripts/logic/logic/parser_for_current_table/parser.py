import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


# first, use "merge and fill.vba"
# to merge and fill cells in the table
# you get "table vba.xlsm"

wb = load_workbook('table merged.xlsx')
sheet = wb.active

# a week has this much days
week = 6

for group_column in range(2, 30 + 1):

    index = ['9:00-10:30', '10:40-12:10', '12:40-14:10', '14:20-15:50',
             '16:00-17:30', '17:40-19:10', '19:20-20:50']
    columns = ['Monday', 'Teacher1', 'Room1',
               'Tuesday', 'Teacher2', 'Room2',
               'Wednesday', 'Teacher3', 'Room3',
               'Thursday', 'Teacher4', 'Room4',
               'Friday', 'Teacher5', 'Room5',
               'Saturday', 'Teacher6', 'Room6']

    df = pd.DataFrame(columns=columns, index=index)
    df.columns.name = 'Time Slots'

    # this much days in a week
    for day in range(week):

        day_row = 3 + day * 22
        day_column = day * 3
        for time_slot in range(7):
            time_slot_row = day_row + 1 + time_slot * 3

            # course name
            df.iloc[time_slot, day_column] = sheet.cell(row=time_slot_row, column=group_column).value
            # teacher
            df.iloc[time_slot, day_column + 1] = sheet.cell(row=time_slot_row + 1, column=group_column).value
            # room
            df.iloc[time_slot, day_column + 2] = sheet.cell(row=time_slot_row + 2, column=group_column).value
    # with pd.option_context('display.max_rows', None, 'display.max_columns', None):  # more options can be specified also
    # print(df)

    group_name = sheet.cell(row=2, column=group_column).value
    year_name = sheet.cell(row=1, column=group_column).value

    groups_data_path = '../../data/groups_schedules'
    if not os.path.exists(f'{groups_data_path}/{year_name}'):
        os.makedirs(f'{groups_data_path}/{year_name}')

    df.to_csv(f'{groups_data_path}/{year_name}/{group_name}.csv')
    # print(f'table-creation scripts/groups_schedules/{sheet.cell(row=2, column=group_column).value}')
