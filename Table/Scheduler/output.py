from openpyxl import load_workbook
from openpyxl.utils.cell import *
from openpyxl.styles import Alignment, PatternFill
from globals import *


def csv_output(results):
    """
    Generates csv file from program output
    :param results: The list of activities to be added to file
    """
    print(cwd)
    with open('Output/Output.csv', 'w') as out:
        out.write("day,slot,targets,a_name(a_type),a_inst,room\n")
        for activity in results:
            out.write(str(activity) + '\n')


def xlsx_output(results):
    """
    Generate xlsx file from program output
    :param results: The list of activities to be added to file
    """
    wb = load_workbook("Output/Template.xlsx")
    ws = wb.active
    for activity in results:
        visualize(ws, activity)
    wb.save("Output/Result.xlsx")


def visualize(ws, activity):
    """
    Given a working sheet and an activity, this function inserts activity data in the appropriate cell(s),
    and formats the cell data (color and alignment)
    :param ws: The working sheet to be operated on
    :param activity: The activity to be positioned
    :return: None
    """
    # Get activity position
    coords = activity_to_index(ws, activity)
    # print(coords)
    for coord in coords:
        # Get the 3 cells involved with activity
        c1 = ws.cell(coord[1], coord[0])
        c2 = ws.cell(coord[1] + 1, coord[0])
        c3 = ws.cell(coord[1] + 2, coord[0])

        # Fill data
        # print("Trying to add " + str(activity) + " into " + str(coord))
        ws[get_column_letter(c1.column) + str(c1.row)] = activity.a_name + ' (' + type_to_str[activity.a_type] + ')'
        ws[get_column_letter(c2.column) + str(c2.row)] = activity.a_inst
        ws[get_column_letter(c3.column) + str(c3.row)] = activity.alloc[2]

        # Text alignment
        c1.alignment = Alignment(horizontal='center')
        c2.alignment = Alignment(horizontal='center')
        c3.alignment = Alignment(horizontal='center')

        # Cell BG color
        c1.fill = PatternFill(fgColor=colors[activity.targets[0][0]], fill_type="solid")
        c2.fill = PatternFill(fgColor=colors[activity.targets[0][0]], fill_type="solid")
        c3.fill = PatternFill(fgColor=colors[activity.targets[0][0]], fill_type="solid")

        # wb.save("Result.xlsx")


def activity_to_index(ws, act):
    """
    Given an activity and working sheet, this function determines the appropriate place(s) for the activity in the
    sheet, merging cells whenever necessary
    :param ws: The working sheet to be operated on
    :param act: The activity to be positioned
    :return: a list of cells coordinates in the form (column, row) for the activity
    """
    row = 3 + 22 * (act.alloc[0] - 1) + 1 + 3 * (act.alloc[1] - 1)

    if len(act.targets) == 1:
        (base, limit) = grp_cnt[(act.targets[0][0], act.targets[0][1])]
        column = base

        # A group of people included, merge from base to end
        if act.targets[0][1] == '**' or act.targets[0][2] == '**':
            for o in range(0, 3):
                s = get_column_letter(base) + str(row + o)
                e = get_column_letter(base + limit - 1) + str(row + o)
                # print("Merging cells from " + s + " to " + e)
                ws.merge_cells(range_string=s + ":" + e)

        else:  # A group is specified
            if int(act.targets[0][2]) > limit:
                raise Exception("No such group number")
            column += int(act.targets[0][2]) - 1
        return [(column, row)]
    else:  # many targets included
        segments = []
        for t in act.targets:
            segments.append((grp_cnt[(t[0], t[1])][0], grp_cnt[(t[0], t[1])][0] + grp_cnt[(t[0], t[1])][1]))

        # Merging Adjacent Segments
        segments.sort()
        # print(segments)
        while True:
            modified = False
            tbr = 0  # to be removed
            for idx in range(len(segments) - 1):
                if segments[idx][1] == segments[idx + 1][0]:
                    segments[idx] = (segments[idx][0], segments[idx + 1][1])
                    tbr = idx + 1
                    modified = True
                    break
            if modified:
                segments.pop(tbr)
            else:
                break
        # print(segments)
        ret = []
        for seg in segments:
            for o in range(0, 3):
                s = get_column_letter(seg[0]) + str(row + o)
                e = get_column_letter(seg[1] - 1) + str(row + o)
                # print("Merging cells from " + s + " to " + e)
                ws.merge_cells(range_string=s + ":" + e)
            ret.append((seg[0], row))
        # column = segments[0][0]
        return ret
