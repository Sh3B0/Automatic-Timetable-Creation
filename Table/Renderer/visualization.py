import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


class Visualization:

    def __init__(self):
        # current working directory
        self.cwd = Path(__file__).parent

        output_path = (self.cwd / '../output/wb.xlsx').resolve()
        if os.path.exists(output_path):
            os.remove(output_path)

        self.workbook = Workbook()

        # these year names will appear in the table
        group_data_path = '../data/groups_schedules'
        years = os.listdir(path=(self.cwd / group_data_path).resolve())

        # these relative file paths are used to fetch groups_schedules for groups_schedules
        groups = [[f'{group_data_path}/{year_dir}/{group_file}'
                   for group_file in os.listdir((self.cwd / f'{group_data_path}/{year_dir}').resolve())]
                  for year_dir in years]

        # there are 3 columns of groups_schedules for each day entry
        self.day_data_width = 3

        # these colors are used for courses
        styles = pd.read_csv((self.cwd / '../data/styles.csv').resolve(),
                             header=None, index_col=0)
        # print(styles)
        day_colors = styles.iloc[0].to_list()
        main_colors = styles.iloc[1].to_list()

        palettes = [{'day': day_color, 'main': main_color} for day_color, main_color in zip(day_colors, main_colors)]
        # print(palettes)
        # exit(0)

        # width of columns
        self.column_width = 32

        # there are this much time slots in a day
        self.time_slots_number = 7

        # apply certain styles for each cell in the table
        self.post_processing()

        # time slot takes this much rows
        self.time_slot_height = 3

        # there will be this much days in a week
        self.week = 6

        # a day takes this much rows
        self.day_height = 1 + self.time_slot_height * self.time_slots_number

        # first create the first column
        self.create_left(groups[0][0], palettes[0])

        # each group takes 1 column
        self.group_width = 1

        # then add courses which start from the 2nd column
        starting_column = 2

        # calculate all starting columns for courses
        course_sizes = [0 if i == 0 else len(groups[i - 1]) for i in range(len(groups))]
        for i in range(1, len(course_sizes)):
            course_sizes[i] = course_sizes[i - 1] + course_sizes[i]
        course_columns = [starting_column + course_size * self.group_width for course_size in course_sizes]

        # add each course
        for year_index in range(len(years)):
            self.add_year(course_name=years[year_index],
                          groups=groups[year_index],
                          course_column=course_columns[year_index],
                          palette=palettes[year_index])

        self.post_processing()

        # save the notebook

        self.workbook.save(filename=(self.cwd / '../output/wb.xlsx').resolve())

    def add_year(self, course_column, course_name, groups, palette):

        # width of a course in columns
        course_width = len(groups) * self.group_width

        # add header of a course
        sheet = self.workbook.active

        # add year_name
        sheet.merge_cells(start_row=1, start_column=course_column,
                          end_row=1, end_column=course_column + course_width - 1)
        sheet.cell(row=1, column=course_column).fill = PatternFill("solid", fgColor=palette['main'])
        sheet.cell(row=1, column=course_column).value = course_name

        # add border around it
        self.add_border(min_row=1, min_column=course_column, max_row=1, max_column=course_column + course_width - 1)

        # first row of each day inside the course should look like merged cells
        for day_index in range(self.week):
            # starting row of a day
            day_row = 3 + day_index * self.day_height
            sheet.merge_cells(start_row=day_row,
                              start_column=course_column,
                              end_row=day_row,
                              end_column=course_column + course_width - 1)
            sheet.cell(row=day_row, column=course_column).fill = \
                PatternFill("solid", fgColor=palette['day'])

        # add groups_schedules
        for group_index in range(len(groups)):
            # all non-empty slots of a course will be painted this color
            group_column = course_column + group_index * self.group_width
            self.add_group(group_column, groups[group_index], palette)

    # adds a column with group groups_schedules into the table
    def add_group(self, group_column, group_data_path, palette):
        wb = self.workbook
        sheet = wb.active

        # 2 first rows are for course name, and group_name
        group_row = 2

        # extract group name from path to group groups_schedules
        group_name = Path(group_data_path).stem

        # put group name into the table
        sheet.merge_cells(start_row=group_row, start_column=group_column,
                          end_row=group_row, end_column=group_column + self.group_width - 1)
        group_cell = sheet.cell(row=group_row, column=group_column)
        group_cell.value = group_name
        group_cell.fill = PatternFill("solid", fgColor=palette['main'])
        self.add_border(min_row=group_row, min_column=group_column, max_row=group_row, max_column=group_column)

        # read groups_schedules for group
        df = pd.read_csv((self.cwd / f'{group_data_path}').resolve(), index_col=0)

        # iterate through df by days
        for day_index in range(self.week):
            # work with subset of dataframe

            # calculate the starting column for a day
            day_column = day_index * self.day_data_width

            # select groups_schedules for one day; no offset since headers are not groups_schedules
            day_data = df.iloc[0: self.time_slots_number, day_column: day_column + self.day_data_width]

            # calculate starting row and column for a day
            day_row = group_row + 1 + self.day_height * day_index
            day_column = group_column

            self.add_day(day_row, day_column, day_data, palette)

    def add_day(self, day_row, day_column, day_data, palette):
        sheet = self.workbook.active

        # iterate through time slots for a day
        for time_slot_index in range(self.time_slots_number):
            # for each day, first row is occupied by an empty row
            time_slot_row = day_row + 1 + time_slot_index * self.time_slot_height

            # put course_name, teacher and room
            # common stands for a lecture or a tutorial

            for index in range(self.time_slot_height):
                row, column = time_slot_row, day_column

                self.add_border(min_row=time_slot_row, min_column=day_column,
                                max_row=time_slot_row + self.time_slot_height - 1,
                                max_column=day_column)

                # no groups_schedules in the dataframe - just merge all cells in the timeslot
                if pd.isnull(day_data.iloc[time_slot_index, index]) and index == 0:
                    sheet.merge_cells(start_row=row, start_column=column,
                                      end_row=row + self.time_slot_height - 1,
                                      end_column=column + self.group_width - 1)
                    break

                cell = sheet.cell(row=row + index, column=column)
                # print(cell.column, cell.row)
                cell.value = day_data.iloc[time_slot_index, index]
                if index == 2 and str(cell.value).find('.') != -1:
                    cell.value = int(cell.value)
                cell.fill = PatternFill("solid", fgColor=palette['main'])

    # creates the leftmost column with names of week days and timeslots
    def create_left(self, path_to_sample_data, palette):
        sheet = self.workbook.active

        # read groups_schedules to get indices
        df = pd.read_csv((self.cwd / path_to_sample_data).resolve(), index_col=0)

        # more options can be specified also
        # with pd.option_context('display.max_rows', None, 'display.max_columns', None):
        #     print(df)

        days = [df.columns[i] for i in range(0, len(df.columns), 3)]
        time_slots = df.index

        # vertical offset
        offset = 3

        # height of a slot in rows
        time_slot_height = self.time_slot_height

        # number of time_slots
        slots_number = len(time_slots)

        # one row for name of day and rows for time_slots
        day_height = 1 + slots_number * time_slot_height
        week = len(days)

        # for each day
        for day_index in range(week):
            # starting on this row
            day_row = offset + day_index * day_height

            # set name of day
            sheet[f'A{day_row}'] = days[day_index]

            # add border
            side_style = Side(style='thin')

            self.add_border(min_row=day_row, min_column=1, max_row=day_row, max_column=1)

            # paint day cell
            sheet.cell(day_row, 1).fill = PatternFill("solid", fgColor=palette['day'])

            # for each time slot
            for time_slot_index in range(slots_number):
                time_slot_row = day_row + 1 + time_slot_index * 3
                sheet.merge_cells(f'A{time_slot_row}:A{time_slot_row + 2}')

                sheet.row_dimensions[time_slot_row].height = 40
                sheet.row_dimensions[time_slot_row + 1].height = 40
                sheet.row_dimensions[time_slot_row + 2].height = 20

                sheet[f'A{time_slot_row}'] = time_slots[time_slot_index]
                sheet[f'A{time_slot_row}'].alignment = Alignment(vertical="top")

                sheet[f'A{time_slot_row}'].border = Border(left=side_style, top=side_style,
                                                           right=side_style, bottom=side_style)

    def post_processing(self):
        # styles applied to all cells
        sheet = self.workbook.active

        # alignment
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text='True')

        # width of columns
        for col in sheet.columns:
            sheet.column_dimensions[get_column_letter(col[0].column)].width = self.column_width

        # borders
        side_style = Side(style='thin')
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if cell.border is None:
                    cell.border = Border(left=side_style, top=side_style, right=side_style, bottom=side_style)

        # freeze first column and 2 top rows
        sheet.freeze_panes = sheet['B3']

        for column in range(1, sheet.max_column):
            for row in range(1, sheet.max_row):
                value = str(sheet.cell(row=row, column=column).value)
                if value.find('Lec') != -1 or value.find('Tut') != -1:
                    nxt = 0
                    while sheet.cell(row=row, column=column + nxt + 1).value == value:
                        nxt += 1
                    sheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + nxt)
                    sheet.merge_cells(start_row=row + 1, start_column=column, end_row=row + 1, end_column=column + nxt)
                    sheet.merge_cells(start_row=row + 2, start_column=column, end_row=row + 2, end_column=column + nxt)

    # add border around some rectangular range of cells
    def add_border(self, min_row, min_column, max_row, max_column):
        sheet = self.workbook.active
        side_style = Side(style='thin')

        for row in range(min_row, max_row + 1):
            sheet.cell(row, min_column).border = Border(left=side_style)
            sheet.cell(row, max_column).border = Border(right=side_style)
        for column in range(min_column, max_column + 1):
            sheet.cell(min_row, column).border = Border(top=side_style)
            sheet.cell(max_row, column).border = Border(bottom=side_style)

        sheet.cell(min_row, min_column).border = Border(left=side_style, top=side_style)
        sheet.cell(min_row, max_column).border = Border(right=side_style, top=side_style)
        sheet.cell(max_row, min_column).border = Border(left=side_style, bottom=side_style)
        sheet.cell(max_row, max_column).border = Border(right=side_style, bottom=side_style)


vis = Visualization()
