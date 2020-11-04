from openpyxl import load_workbook
from openpyxl.utils.cell import *
from openpyxl.styles import Alignment, PatternFill
from globals import *


def csvOutput(results):
    with open('out/Output.csv', 'w') as out:
        out.write("day,slot,target,a_name,a_inst(a_type),room\n")
        for activity in results:
            out.write(str(activity) + '\n')


def xlsxOutput(results):
    wb = load_workbook(filename="out/Template.xlsx")
    ws = wb.active
    for activity in results:
        visualize(ws, activity)
    wb.save("out/Result.xlsx")


def visualize(ws, data):
    # Get activity position
    coord = ActivityToIndex(ws, data)
    # print(coord)

    # Get the 3 cells involved with activity
    c1 = ws.cell(coord[1], coord[0])
    c2 = ws.cell(coord[1] + 1, coord[0])
    c3 = ws.cell(coord[1] + 2, coord[0])

    # Fill data
    # print("Trying to add " + str(data) + " into " + str(coord))
    ws[get_column_letter(c1.column) + str(c1.row)] = data.a_name + ' (' + type_to_str[data.a_type] + ')'
    ws[get_column_letter(c2.column) + str(c2.row)] = data.a_inst
    ws[get_column_letter(c3.column) + str(c3.row)] = data.alloc[2]

    # Text alignment
    c1.alignment = Alignment(horizontal='center')
    c2.alignment = Alignment(horizontal='center')
    c3.alignment = Alignment(horizontal='center')

    # Cell BG color
    c1.fill = PatternFill(fgColor=colors[data.target[0]], fill_type="solid")
    c2.fill = PatternFill(fgColor=colors[data.target[0]], fill_type="solid")
    c3.fill = PatternFill(fgColor=colors[data.target[0]], fill_type="solid")

    # wb.save("Result.xlsx")


def ActivityToIndex(ws, act):
    row = 3 + 22 * (act.alloc[0] - 1) + 1 + 3 * (act.alloc[1] - 1)
    (base, limit) = grp_cnt[(act.target[0], act.target[1])]
    column = base

    # A group of people included, merge from base to base + limit - 1
    if act.target[1] == '**' or act.target[2] == '**':
        for o in range(0, 3):
            s = get_column_letter(base) + str(row + o)
            e = get_column_letter(base + limit - 1) + str(row + o)
            # print("Merging cells from " + s + " to " + e)
            ws.merge_cells(range_string=s + ":" + e)

    else:  # A group is specified
        if int(act.target[2]) > limit:
            raise Exception("No such group number")
        column += int(act.target[2]) - 1

    return column, row
